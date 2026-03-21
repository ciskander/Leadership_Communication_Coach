#!/usr/bin/env python3
"""
view_analysis.py — Parse model output (mvp.v0.3.0) and produce a formatted
Excel workbook for evaluating taxonomy v3.0 pattern scoring at a glance.

Usage:
    python scripts/view_analysis.py output1.json [output2.json ...] -o report.xlsx
    cat output.json | python scripts/view_analysis.py - -o report.xlsx
"""

import argparse
import glob
import json
import os
import sys
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

PATTERNS = [
    ("purposeful_framing", "Purposeful Framing", "Meeting Structure", "dual_element"),
    ("focus_management", "Focus Management", "Meeting Structure", "tiered_rubric"),
    ("participation_management", "Participation Mgmt", "Participation", "tiered_rubric"),
    ("disagreement_navigation", "Disagreement Nav", "Participation", "tiered_rubric"),
    ("resolution_and_alignment", "Resolution & Alignment", "Decisions", "dual_element"),
    ("assignment_clarity", "Assignment Clarity", "Decisions", "complexity_tiered"),
    ("question_quality", "Question Quality", "Communication", "binary"),
    ("communication_clarity", "Communication Clarity", "Communication", "tiered_rubric"),
    ("feedback_quality", "Feedback Quality", "Communication", "multi_element"),
]

PATTERN_IDS = [p[0] for p in PATTERNS]
PATTERN_NAMES = {p[0]: p[1] for p in PATTERNS}
PATTERN_CLUSTERS = {p[0]: p[2] for p in PATTERNS}
PATTERN_SCORING = {p[0]: p[3] for p in PATTERNS}

# Fills
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GRAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DELTA_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
DELTA_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
EXCLUDED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# Fonts
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD_FONT = Font(bold=True, size=11)
NORMAL_FONT = Font(size=11)
DIM_FONT = Font(color="999999", size=11)
SMALL_FONT = Font(size=10)
SMALL_DIM_FONT = Font(size=10, color="999999")

# Borders
THIN_BORDER = Border(
    bottom=Side(style="thin", color="B4B4B4"),
)

# Alignment
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top")
CENTER_WRAP = Alignment(horizontal="center", vertical="top", wrap_text=True)


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------


def load_analysis(path: str) -> dict[str, Any]:
    """Load a JSON analysis file. Use '-' for stdin."""
    if path == "-":
        raw = sys.stdin.read()
    else:
        with open(path, "r") as f:
            raw = f.read()
    data = json.loads(raw)
    if data.get("schema_version") != "mvp.v0.3.0":
        print(
            f"Warning: {path} has schema_version "
            f"'{data.get('schema_version')}', expected 'mvp.v0.3.0'",
            file=sys.stderr,
        )
    return data


def build_span_index(
    data: dict,
) -> tuple[dict[str, dict], dict[tuple[int, int], dict]]:
    """Build lookup dicts for evidence spans — by ID and by turn range."""
    spans = data.get("evidence_spans", [])
    by_id: dict[str, dict] = {}
    by_turns: dict[tuple[int, int], dict] = {}
    for s in spans:
        sid = s.get("evidence_span_id", "")
        by_id[sid] = s
        start = s.get("turn_start_id")
        end = s.get("turn_end_id")
        if start is not None and end is not None:
            by_turns[(start, end)] = s
    return by_id, by_turns


def get_analysis_id(data: dict) -> str:
    """Return a short label for the run — prefer meeting_id over analysis_id."""
    mid = data.get("context", {}).get("meeting_id")
    if mid:
        return str(mid)
    return data.get("meta", {}).get("analysis_id", "unknown")


def get_pattern_snapshot_map(data: dict) -> dict[str, dict]:
    """Map pattern_id -> snapshot item for quick lookup."""
    result: dict[str, dict] = {}
    for item in data.get("pattern_snapshot", []):
        pid = item.get("pattern_id")
        if pid:
            result[pid] = item
    return result


def score_fill(score: float | None) -> PatternFill:
    if score is None:
        return GRAY_FILL
    if score >= 0.75:
        return GREEN_FILL
    if score >= 0.5:
        return YELLOW_FILL
    return RED_FILL


def build_back_references(data: dict) -> dict[str, list[str]]:
    """Map span_id -> list of pattern_ids that reference it."""
    refs: dict[str, list[str]] = {}
    for item in data.get("pattern_snapshot", []):
        pid = item.get("pattern_id", "")
        for sid in item.get("evidence_span_ids", []):
            refs.setdefault(sid, []).append(pid)
    return refs


def set_col_widths(ws, widths: dict[int, float]):
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def write_header_row(ws, row: int, values: list[str], col_start: int = 1):
    for i, val in enumerate(values):
        cell = ws.cell(row=row, column=col_start + i, value=val)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_WRAP


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def build_summary_sheet(
    wb: Workbook,
    analyses: list[dict],
    pattern_filter: list[str] | None,
    evaluable_only: bool,
):
    ws = wb.active
    ws.title = "Summary"

    # -- Context header rows --
    row = 1
    for idx, data in enumerate(analyses):
        aid = get_analysis_id(data)
        ctx = data.get("context", {})
        meta = data.get("meta", {})
        speaker = ctx.get("target_speaker_label", "?")
        role = ctx.get("target_role", "?")
        mtype = ctx.get("meeting_type", "?")
        mid = ctx.get("meeting_id", "?")
        atype = meta.get("analysis_type", "?")

        label = f"{aid} | {atype} | Speaker: {speaker} ({role}) | {mtype} | {mid}"
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = BOLD_FONT
        row += 1

    row += 1  # blank row

    # -- Column structure --
    # Fixed cols: Pattern, Cluster, Scoring Type
    # Per analysis: Status, Score, Opps, Detail
    fixed_cols = ["Pattern", "Cluster", "Scoring Type"]
    per_analysis_cols = ["Status", "Score", "Opps", "Detail"]
    n_analyses = len(analyses)

    headers = list(fixed_cols)
    for data in analyses:
        aid = get_analysis_id(data)
        prefix = aid if n_analyses > 1 else ""
        for col_name in per_analysis_cols:
            headers.append(f"{prefix} {col_name}".strip() if prefix else col_name)

    # Delta columns if 2+ analyses
    if n_analyses >= 2:
        headers.append("Delta (last - first)")

    write_header_row(ws, row, headers)
    header_row = row
    row += 1

    # -- Pattern rows --
    patterns_to_show = PATTERN_IDS
    if pattern_filter:
        patterns_to_show = [p for p in PATTERN_IDS if p in pattern_filter]

    snapshot_maps = [get_pattern_snapshot_map(d) for d in analyses]

    for pid in patterns_to_show:
        # Check evaluable_only filter
        if evaluable_only:
            any_evaluable = any(
                sm.get(pid, {}).get("evaluable_status") == "evaluable"
                for sm in snapshot_maps
            )
            if not any_evaluable:
                continue

        col = 1
        ws.cell(row=row, column=col, value=PATTERN_NAMES.get(pid, pid)).font = (
            BOLD_FONT
        )
        col += 1
        ws.cell(
            row=row, column=col, value=PATTERN_CLUSTERS.get(pid, "")
        ).font = NORMAL_FONT
        col += 1
        ws.cell(
            row=row, column=col, value=PATTERN_SCORING.get(pid, "")
        ).font = NORMAL_FONT
        col += 1

        first_score = None
        last_score = None

        for i, sm in enumerate(snapshot_maps):
            item = sm.get(pid, {})
            status = item.get("evaluable_status", "—")
            score = item.get("score")
            opps = item.get("opportunity_count")

            # Status
            status_cell = ws.cell(row=row, column=col, value=status)
            status_cell.alignment = CENTER
            if status == "insufficient_signal":
                status_cell.font = Font(color="B8860B", size=11)
            elif status == "not_evaluable":
                status_cell.font = DIM_FONT
            else:
                status_cell.font = NORMAL_FONT
            col += 1

            # Score
            if score is not None:
                score_cell = ws.cell(row=row, column=col, value=round(score, 2))
                score_cell.fill = score_fill(score)
                score_cell.alignment = CENTER
                score_cell.number_format = "0.00"
                if first_score is None:
                    first_score = score
                last_score = score
            else:
                score_cell = ws.cell(row=row, column=col, value="—")
                score_cell.font = DIM_FONT
                score_cell.fill = GRAY_FILL
                score_cell.alignment = CENTER
            col += 1

            # Opps
            if opps is not None:
                ws.cell(row=row, column=col, value=opps).alignment = CENTER
            else:
                c = ws.cell(row=row, column=col, value="—")
                c.font = DIM_FONT
                c.alignment = CENTER
            col += 1

            # Detail (pattern-specific)
            detail = _build_detail(pid, item)
            detail_cell = ws.cell(row=row, column=col, value=detail)
            detail_cell.font = SMALL_FONT if detail else DIM_FONT
            detail_cell.alignment = CENTER
            col += 1

        # Delta column
        if n_analyses >= 2:
            if first_score is not None and last_score is not None:
                delta = last_score - first_score
                delta_cell = ws.cell(
                    row=row, column=col, value=round(delta, 2)
                )
                delta_cell.number_format = "+0.00;-0.00;0.00"
                delta_cell.alignment = CENTER
                if delta > 0.005:
                    delta_cell.fill = DELTA_GREEN_FILL
                elif delta < -0.005:
                    delta_cell.fill = DELTA_RED_FILL
            else:
                c = ws.cell(row=row, column=col, value="n/a")
                c.font = DIM_FONT
                c.alignment = CENTER

        # Bottom border for row
        for c in range(1, col + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER

        row += 1

    # -- Column widths --
    widths = {1: 24, 2: 18, 3: 16}
    col = 4
    for _ in analyses:
        widths[col] = 18  # status
        widths[col + 1] = 8  # score
        widths[col + 2] = 6  # opps
        widths[col + 3] = 16  # detail
        col += 4
    if n_analyses >= 2:
        widths[col] = 14
    set_col_widths(ws, widths)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=4)


def _build_detail(pid: str, item: dict) -> str:
    """Build pattern-specific detail string."""
    scoring = PATTERN_SCORING.get(pid, "")
    parts = []
    if scoring == "dual_element":
        a = item.get("element_a_count")
        b = item.get("element_b_count")
        if a is not None:
            parts.append(f"A:{a}")
        if b is not None:
            parts.append(f"B:{b}")
    elif scoring == "complexity_tiered":
        s = item.get("simple_count")
        c = item.get("complex_count")
        if s is not None:
            parts.append(f"S:{s}")
        if c is not None:
            parts.append(f"C:{c}")
    if pid == "participation_management":
        ba = item.get("balance_assessment")
        if ba:
            parts.append(ba)
    return " ".join(parts)


def build_events_sheet(
    wb: Workbook,
    data: dict,
    analysis_id: str,
    span_by_id: dict,
    span_by_turns: dict,
    pattern_filter: list[str] | None,
    evaluable_only: bool,
):
    sheet_name = f"Events — {analysis_id}"[:31]  # Excel 31-char limit
    ws = wb.create_sheet(title=sheet_name)

    headers = [
        "Pattern",
        "Event ID",
        "Turns",
        "Control",
        "Decision",
        "Success",
        "Reason Code",
        "Notes",
        "Evidence Excerpt",
    ]
    write_header_row(ws, 1, headers)
    row = 2

    snapshot = data.get("pattern_snapshot", [])
    has_any_events = False

    for item in snapshot:
        pid = item.get("pattern_id", "")
        if pattern_filter and pid not in pattern_filter:
            continue
        if evaluable_only and item.get("evaluable_status") != "evaluable":
            continue

        events = item.get("opportunity_events", [])
        if not events:
            continue

        has_any_events = True
        pattern_name = PATTERN_NAMES.get(pid, pid)
        score = item.get("score")
        opps = item.get("opportunity_count")

        # Pattern header row
        label = f"{pattern_name} — score: {score}, opps: {opps}"
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = BOLD_FONT
        cell.fill = LIGHT_BLUE_FILL
        for c in range(2, len(headers) + 1):
            ws.cell(row=row, column=c).fill = LIGHT_BLUE_FILL
        row += 1

        for ev in events:
            event_id = ev.get("event_id", "")
            t_start = ev.get("turn_start_id", "?")
            t_end = ev.get("turn_end_id", "?")
            control = ev.get("target_control", "")
            decision = ev.get("count_decision", "")
            success = ev.get("success")
            reason = ev.get("reason_code", "")
            notes = ev.get("notes", "") or ""
            is_excluded = decision == "excluded"

            # Look up evidence excerpt
            excerpt = _find_excerpt(t_start, t_end, span_by_id, span_by_turns, item)

            col = 1
            ws.cell(row=row, column=col, value=pattern_name)
            col += 1
            ws.cell(row=row, column=col, value=event_id)
            col += 1
            ws.cell(
                row=row, column=col, value=f"{t_start}–{t_end}"
            ).alignment = CENTER
            col += 1
            ws.cell(row=row, column=col, value=control).alignment = CENTER
            col += 1
            ws.cell(row=row, column=col, value=decision).alignment = CENTER
            col += 1

            # Success cell with color
            if success is not None:
                sc = ws.cell(row=row, column=col, value=round(success, 2))
                sc.number_format = "0.00"
                sc.alignment = CENTER
                if not is_excluded:
                    sc.fill = score_fill(success)
            else:
                ws.cell(row=row, column=col, value="—").alignment = CENTER
            col += 1

            ws.cell(row=row, column=col, value=reason)
            col += 1
            ws.cell(row=row, column=col, value=notes).alignment = WRAP
            col += 1

            # Excerpt (truncated)
            if excerpt:
                display = excerpt[:200] + ("..." if len(excerpt) > 200 else "")
            else:
                display = "(no matching span)"
            ws.cell(row=row, column=col, value=display).alignment = WRAP

            # Dim excluded rows
            if is_excluded:
                for c in range(1, len(headers) + 1):
                    cell = ws.cell(row=row, column=c)
                    cell.font = SMALL_DIM_FONT
                    cell.fill = EXCLUDED_FILL

            row += 1

        # Blank separator row between patterns
        row += 1

    if not has_any_events:
        ws.cell(
            row=2, column=1, value="No opportunity events found in this analysis."
        ).font = DIM_FONT

    set_col_widths(
        ws, {1: 22, 2: 9, 3: 9, 4: 9, 5: 10, 6: 9, 7: 26, 8: 30, 9: 60}
    )
    ws.freeze_panes = "A2"


def _find_excerpt(
    t_start: int,
    t_end: int,
    span_by_id: dict,
    span_by_turns: dict,
    pattern_item: dict,
) -> str:
    """Find the evidence excerpt for an opportunity event's turn range."""
    # Direct turn match
    span = span_by_turns.get((t_start, t_end))
    if span:
        return span.get("excerpt", "")

    # Fallback: check if any evidence span overlaps the turn range
    for (s, e), sp in span_by_turns.items():
        if s <= t_start and e >= t_end:
            return sp.get("excerpt", "")
        if s >= t_start and s <= t_end:
            return sp.get("excerpt", "")

    return ""


def build_coaching_sheet(wb: Workbook, data: dict, analysis_id: str):
    sheet_name = f"Coaching — {analysis_id}"[:31]
    ws = wb.create_sheet(title=sheet_name)

    write_header_row(ws, 1, ["Section", "Pattern", "Content"])
    row = 2

    coaching = data.get("coaching_output", {})

    # Strengths
    for i, s in enumerate(coaching.get("strengths", [])):
        ws.cell(row=row, column=1, value=f"Strength {i + 1}").font = BOLD_FONT
        ws.cell(row=row, column=2, value=s.get("pattern_id", ""))
        ws.cell(row=row, column=3, value=s.get("message", "")).alignment = WRAP
        row += 1

    # Focus
    for f in coaching.get("focus", []):
        ws.cell(row=row, column=1, value="Focus").font = BOLD_FONT
        ws.cell(row=row, column=2, value=f.get("pattern_id", ""))
        ws.cell(row=row, column=3, value=f.get("message", "")).alignment = WRAP
        row += 1

    # Micro-experiment
    for exp in coaching.get("micro_experiment", []):
        row += 1  # blank separator
        ws.cell(row=row, column=1, value="Experiment").font = BOLD_FONT
        eid = exp.get("experiment_id", "")
        title = exp.get("title", "")
        ws.cell(row=row, column=2, value=exp.get("pattern_id", ""))
        ws.cell(row=row, column=3, value=f'{eid}: "{title}"').alignment = WRAP
        row += 1

        ws.cell(row=row, column=1, value="Instruction").font = NORMAL_FONT
        ws.cell(
            row=row, column=3, value=exp.get("instruction", "")
        ).alignment = WRAP
        row += 1

        ws.cell(row=row, column=1, value="Success Marker").font = NORMAL_FONT
        ws.cell(
            row=row, column=3, value=exp.get("success_marker", "")
        ).alignment = WRAP
        row += 1

    # Experiment tracking
    tracking = data.get("experiment_tracking", {})
    detection = tracking.get("detection_in_this_meeting")
    if detection:
        row += 1
        ws.cell(row=row, column=1, value="Detection").font = BOLD_FONT
        attempt = detection.get("attempt", "")
        count = detection.get("count_attempts", 0)
        ws.cell(
            row=row,
            column=3,
            value=f"Attempt: {attempt} | Count: {count}",
        ).alignment = WRAP
        row += 1

        cn = detection.get("coaching_note")
        if cn:
            ws.cell(row=row, column=1, value="Detection Note")
            ws.cell(row=row, column=3, value=cn).alignment = WRAP
            row += 1

    set_col_widths(ws, {1: 16, 2: 24, 3: 80})
    ws.freeze_panes = "A2"


def build_evidence_sheet(
    wb: Workbook, data: dict, analysis_id: str, back_refs: dict
):
    sheet_name = f"Evidence — {analysis_id}"[:31]
    ws = wb.create_sheet(title=sheet_name)

    write_header_row(
        ws, 1, ["Span ID", "Turns", "Speaker Role", "Excerpt", "Referenced By"]
    )
    row = 2

    for span in data.get("evidence_spans", []):
        sid = span.get("evidence_span_id", "")
        t_start = span.get("turn_start_id", "?")
        t_end = span.get("turn_end_id", "?")
        role = span.get("speaker_role", "")
        excerpt = span.get("excerpt", "")
        refs = back_refs.get(sid, [])

        ws.cell(row=row, column=1, value=sid).font = BOLD_FONT
        ws.cell(
            row=row, column=2, value=f"{t_start}–{t_end}"
        ).alignment = CENTER
        ws.cell(row=row, column=3, value=role or "—").alignment = CENTER
        ws.cell(row=row, column=4, value=excerpt).alignment = WRAP
        ws.cell(
            row=row,
            column=5,
            value=", ".join(PATTERN_NAMES.get(r, r) for r in refs),
        ).alignment = WRAP
        ws.cell(row=row, column=1).border = THIN_BORDER
        row += 1

    set_col_widths(ws, {1: 10, 2: 9, 3: 14, 4: 70, 5: 35})
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Parse model analysis output (mvp.v0.3.0) into a formatted Excel workbook.",
    )
    parser.add_argument(
        "files",
        nargs="*",
        metavar="FILE",
        help="JSON analysis file(s). Use - for stdin.",
    )
    parser.add_argument(
        "-d",
        "--dir",
        metavar="DIR",
        help="Process all .json files in this directory",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="analysis_report.xlsx",
        help="Output .xlsx path (default: analysis_report.xlsx)",
    )
    parser.add_argument(
        "-p",
        "--pattern",
        action="append",
        dest="patterns",
        metavar="ID",
        help="Filter to specific pattern_id (repeatable)",
    )
    parser.add_argument(
        "-e",
        "--evaluable-only",
        action="store_true",
        help="Only include evaluable patterns",
    )
    parser.add_argument(
        "--no-events",
        action="store_true",
        help="Omit opportunity events detail sheets",
    )

    args = parser.parse_args()

    # Validate pattern filters
    if args.patterns:
        for p in args.patterns:
            if p not in PATTERN_IDS:
                parser.error(
                    f"Unknown pattern '{p}'. Valid: {', '.join(PATTERN_IDS)}"
                )

    # Collect file paths
    file_paths = list(args.files or [])
    if args.dir:
        dir_path = os.path.expanduser(args.dir)
        if not os.path.isdir(dir_path):
            parser.error(f"Not a directory: {dir_path}")
        found = sorted(glob.glob(os.path.join(dir_path, "*.json")))
        if not found:
            parser.error(f"No .json files found in {dir_path}")
        file_paths.extend(found)

    if not file_paths:
        parser.error("No input files. Provide FILE arguments or use --dir.")

    # Load analyses
    analyses = []
    for path in file_paths:
        try:
            analyses.append(load_analysis(path))
        except (json.JSONDecodeError, FileNotFoundError) as exc:
            print(f"Error loading {path}: {exc}", file=sys.stderr)
            sys.exit(1)

    if not analyses:
        print("No analysis files loaded.", file=sys.stderr)
        sys.exit(1)

    # Build workbook
    wb = Workbook()

    # Summary sheet (shared across all analyses)
    build_summary_sheet(wb, analyses, args.patterns, args.evaluable_only)

    # Per-analysis detail sheets
    for data in analyses:
        aid = get_analysis_id(data)
        span_by_id, span_by_turns = build_span_index(data)
        back_refs = build_back_references(data)

        if not args.no_events:
            build_events_sheet(
                wb,
                data,
                aid,
                span_by_id,
                span_by_turns,
                args.patterns,
                args.evaluable_only,
            )

        build_coaching_sheet(wb, data, aid)
        build_evidence_sheet(wb, data, aid, back_refs)

    # Save
    wb.save(args.output)
    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()
